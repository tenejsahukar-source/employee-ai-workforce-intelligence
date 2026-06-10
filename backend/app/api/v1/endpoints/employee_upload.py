"""
backend/app/routes/employee_upload.py

FastAPI router for employee list (paginated) and CSV/XLSX bulk upload.
Upload jobs are tracked in a simple in-memory dict for demo purposes;
swap for Redis + Celery in production.
"""

from __future__ import annotations

import asyncio
import csv
import io
import logging
import uuid
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import Employee
from app.schemas import (
    EmployeeListItem,
    EmployeeListResponse,
    BulkUploadResponse,
    UploadProgressResponse,
    UploadValidationError,
)
from app.auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/employees", tags=["employees"])

# ── In-memory job store (replace with Redis in production) ────────────────────
_JOBS: Dict[str, dict] = {}


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/v1/employees
# ─────────────────────────────────────────────────────────────────────────────

@router.get("", response_model=EmployeeListResponse)
async def list_employees(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    dept: Optional[str] = Query(default=None),
    search: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Paginated employee list with optional department and search filters."""
    logger.info(f"[employees.list] page={page} page_size={page_size} dept={dept} search={search}")

    stmt = select(Employee)

    if dept:
        stmt = stmt.where(Employee.dept.ilike(f"%{dept}%"))
    if search:
        stmt = stmt.where(
            Employee.name.ilike(f"%{search}%") |
            Employee.email.ilike(f"%{search}%") |
            Employee.role.ilike(f"%{search}%")
        )

    # Total count
    count_result = await db.execute(
        select(Employee).where(stmt.whereclause) if stmt.whereclause is not None else select(Employee)
    )
    total = len(count_result.scalars().all())

    # Paginated results
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(stmt)
    employees_db = result.scalars().all()

    return EmployeeListResponse(
        employees=[_to_list_item(e) for e in employees_db],
        total=total,
        page=page,
        page_size=page_size,
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/v1/employees/:id
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{employee_id}", response_model=EmployeeListItem)
async def get_employee(
    employee_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    result = await db.execute(select(Employee).where(Employee.id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    return _to_list_item(emp)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/v1/employees/upload
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=BulkUploadResponse)
async def upload_employees(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Accept CSV or XLSX file, validate rows, bulk-insert employees.
    Returns a job_id for progress polling.
    """
    logger.info(f"[employees.upload] filename={file.filename} content_type={file.content_type}")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"csv", "xlsx", "xls"}:
        raise HTTPException(status_code=422, detail=f"Unsupported file type: .{ext}")

    content = await file.read()
    job_id  = str(uuid.uuid4())

    _JOBS[job_id] = {"status": "processing", "progress": 0, "inserted": 0, "total_rows": 0, "errors": []}

    # Process synchronously for simplicity; move to Celery task in production
    try:
        rows = _parse_file(content, ext, job_id)
    except Exception as exc:
        logger.error(f"[employees.upload] Parse error: {exc}")
        _JOBS[job_id]["status"] = "failed"
        raise HTTPException(status_code=422, detail=str(exc))

    inserted, updated, skipped, errors = await _bulk_upsert(rows, db, job_id)

    _JOBS[job_id]["status"] = "done"
    _JOBS[job_id]["progress"] = 100

    return BulkUploadResponse(
        inserted=inserted,
        updated=updated,
        skipped=skipped,
        errors=errors,
        job_id=job_id,
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/v1/employees/upload/progress/:job_id
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/upload/progress/{job_id}", response_model=UploadProgressResponse)
async def upload_progress(
    job_id: str,
    current_user=Depends(get_current_user),
):
    job = _JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    return UploadProgressResponse(
        job_id=job_id,
        status=job["status"],
        progress=job["progress"],
        inserted=job["inserted"],
        total_rows=job["total_rows"],
        errors=job["errors"],
    )


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

REQUIRED_COLUMNS = {
    "name", "role", "dept", "email",
}

def _parse_file(content: bytes, ext: str, job_id: str) -> List[dict]:
    """Parse CSV or XLSX bytes into a list of row dicts."""
    if ext == "csv":
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
        rows = list(reader)
    else:
        # xlsx / xls — requires openpyxl
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for XLSX uploads: pip install openpyxl")

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})

    if not rows:
        raise ValueError("File is empty or has no data rows")

    _JOBS[job_id]["total_rows"] = len(rows)
    logger.info(f"[employees.upload] Parsed {len(rows)} rows from {ext}")
    return rows


async def _bulk_upsert(
    rows: List[dict],
    db: AsyncSession,
    job_id: str,
) -> tuple[int, int, int, List[UploadValidationError]]:
    """Validate and upsert each row. Returns (inserted, updated, skipped, errors)."""
    inserted = updated = skipped = 0
    errors: List[UploadValidationError] = []

    for i, row in enumerate(rows, start=2):  # row 1 = headers
        row_lower = {k.lower().strip().replace(" ", "_"): v for k, v in row.items()}

        # Validate required columns
        row_errors = []
        for col in REQUIRED_COLUMNS:
            if not row_lower.get(col):
                row_errors.append(UploadValidationError(
                    row=i,
                    field=col,
                    message=f"Missing required field: '{col}'"
                ))

        if row_errors:
            errors.extend(row_errors)
            skipped += 1
            continue

        # Check for existing employee by email
        existing_result = await db.execute(
            select(Employee).where(Employee.email == row_lower["email"])
        )
        existing = existing_result.scalar_one_or_none()

        try:
            if existing:
                # Update mutable fields
                existing.name   = row_lower.get("name", existing.name)
                existing.role   = row_lower.get("role", existing.role)
                existing.dept   = row_lower.get("dept", existing.dept)
                existing.manager = row_lower.get("manager", existing.manager)
                existing.attrition_risk = float(row_lower.get("attrition_risk", existing.attrition_risk or 0))
                updated += 1
            else:
                emp = Employee(
                    id=str(uuid.uuid4()),
                    name=row_lower["name"],
                    role=row_lower.get("role", "Employee"),
                    dept=row_lower.get("dept", "Unknown"),
                    email=row_lower["email"],
                    manager=row_lower.get("manager", ""),
                    manager_role=row_lower.get("manager_role", row_lower.get("managerrole", "")),
                    tenure=row_lower.get("tenure", ""),
                    attrition_risk=float(row_lower.get("attrition_risk", 0)),
                    age=int(row_lower["age"]) if row_lower.get("age") else None,
                    years_at_company=int(row_lower.get("years_at_company", 0) or 0),
                    department=row_lower.get("department", row_lower.get("dept", "")),
                    over_time=row_lower.get("over_time", row_lower.get("overtime", "No")),
                    monthly_income=float(row_lower.get("monthly_income", 0) or 0),
                    job_level=int(row_lower.get("job_level", 1) or 1),
                    job_satisfaction=int(row_lower.get("job_satisfaction", 3) or 3),
                    work_life_balance=int(row_lower.get("work_life_balance", 3) or 3),
                )
                db.add(emp)
                inserted += 1

        except (ValueError, TypeError) as exc:
            errors.append(UploadValidationError(
                row=i,
                field="data",
                message=f"Type conversion error: {exc}"
            ))
            skipped += 1
            continue

        # Update progress every 50 rows
        if i % 50 == 0:
            _JOBS[job_id]["progress"] = min(95, round((i / len(rows)) * 100))
            _JOBS[job_id]["inserted"] = inserted
            _JOBS[job_id]["errors"]   = errors
            await db.flush()

    await db.commit()
    _JOBS[job_id]["inserted"] = inserted
    return inserted, updated, skipped, errors


def _to_list_item(emp: Employee) -> EmployeeListItem:
    """Convert SQLAlchemy Employee model to Pydantic EmployeeListItem."""
    return EmployeeListItem(
        id=str(emp.id),
        name=emp.name or "",
        role=emp.role or "",
        dept=emp.dept or "",
        tenure=str(emp.tenure or ""),
        risk=float(emp.attrition_risk or 0),
        email=emp.email or "",
        manager=emp.manager or "",
        manager_role=getattr(emp, "manager_role", ""),
        skills=getattr(emp, "skills", []) or [],
        certifications=getattr(emp, "certifications", []) or [],
        work_location=getattr(emp, "work_location", None),
        employment_type=getattr(emp, "employment_type", None),
        status=getattr(emp, "status", "Active"),
        performance_score=getattr(emp, "performance_score", None),
        satisfaction_level=getattr(emp, "satisfaction_level", None),
        salary_band=getattr(emp, "salary_band", None),
        age=getattr(emp, "age", None),
        gender=getattr(emp, "gender", None),
        education=getattr(emp, "education", None),
        marital_status=getattr(emp, "marital_status", None),
        years_in_role=str(getattr(emp, "years_in_role", "") or ""),
        overtime_status=getattr(emp, "over_time", None),
    )
